from atm_management.generic_view_mixins.custom_api_view import CustomAPIView
from rest_framework.response import Response
from atm_management.serializers import UploadFileSerializer, CityDataSerializer
from atm_management.exceptions.rest_api_exception import RestAPIException
from openpyxl import load_workbook
from io import BytesIO
from atm_management.models import ATMSite, ATMCity, ATMCity
from atm_management.model_queries import ATMSiteModelQueries, ATMCityModelQueries, ATMCityModelQueries
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework import generics  


COLUMN_HEADERS = [
    "City",
    "State"
]
COULMNS_LEN = 2

FIELD_COULMN_MAPPING = {
    "City": "name",
    "State": "state",
}


class UploadCityView(CustomAPIView, ATMCityModelQueries):
    
    serializer_class = UploadFileSerializer

    def get_column_mapping(self, row):
        mapping_dict = dict()
        for index, cell in enumerate(row):
            if cell.value in COLUMN_HEADERS:
                mapping_dict[FIELD_COULMN_MAPPING[cell.value]] = index      
            else:
                raise RestAPIException(f"Incorrect Column Name {cell.value}")

        if len(mapping_dict) < COULMNS_LEN:
            raise RestAPIException(f"Missing Columns from : {COLUMN_HEADERS}")
        return mapping_dict
    
    def preprare_serializer_data(self, worksheet, mapping_dict):
        data_list = list()
        for row in worksheet.iter_rows(min_row=2,max_col=COULMNS_LEN):
            if any(cell.value for cell in row):
                serializer_dict = dict()
                for field in mapping_dict:
                    serializer_dict[field] = row[mapping_dict[field]].value
                data_list.append(serializer_dict)
        return data_list

    def post(self, request):
        result = list()
        req_valid_data = self.validate_serializer(UploadFileSerializer, request.data)      
        workbook = load_workbook(req_valid_data["file"])
        worksheet = workbook.active

        mapping = self.get_column_mapping(worksheet[1])
        data_list = self.preprare_serializer_data(worksheet, mapping)
        record_valid_data = self.validate_many_serializer(CityDataSerializer, data_list)
        for data in record_valid_data:
            print(data)
            created, city_obj = self.update_or_create_city(city=data["name"], state=data["state"])
            result.append({"id": city_obj.id, "name": city_obj.name.title(), "created": created if created else False, "modifed": True if not created else False})
        return Response({"status": 1, "result": result})
    

   

    