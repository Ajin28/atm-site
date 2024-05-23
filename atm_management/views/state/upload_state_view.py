from atm_management.generic_view_mixins.custom_api_view import CustomAPIView
from rest_framework.response import Response
from atm_management.serializers import UploadFileSerializer, StateDataSerializer
from atm_management.exceptions.rest_api_exception import RestAPIException
from openpyxl import load_workbook
from io import BytesIO
from atm_management.models import ATMSite, ATMCity, ATMState
from atm_management.model_queries import ATMSiteModelQueries, ATMStateModelQueries, ATMCityModelQueries
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework import generics  


COLUMN_HEADERS = [
    "State",
    "Abbreviation"
]
COULMNS_LEN = 2

FIELD_COULMN_MAPPING = {
    "Abbreviation": "abbreviation",
    "State": "name",
}


class UploadStateView(CustomAPIView, ATMSiteModelQueries, ATMCityModelQueries, ATMStateModelQueries):
    
    serializer_class = UploadFileSerializer

    def get_column_mapping(self, row):
        mapping_dict = dict()
        for index, cell in enumerate(row):
            if cell.value in COLUMN_HEADERS:
                mapping_dict[FIELD_COULMN_MAPPING[cell.value]] = index      
            else:
                raise RestAPIException(f"Incorrect Column Name {cell.value}")

        if len(mapping_dict) < COULMNS_LEN:
            raise RestAPIException(f"Missing Columns from : {COLUMN_HEADERS}")
        return mapping_dict
    
    def preprare_serializer_data(self, worksheet, mapping_dict):
        data_list = list()
        for row in worksheet.iter_rows(min_row=2,max_col=COULMNS_LEN):
            if any(cell.value for cell in row):
                serializer_dict = dict()
                for field in mapping_dict:
                    serializer_dict[field] = row[mapping_dict[field]].value
                data_list.append(serializer_dict)
        return data_list

    def post(self, request):
        result = list()
        req_valid_data = self.validate_serializer(UploadFileSerializer, request.data)      
        workbook = load_workbook(req_valid_data["file"])
        worksheet = workbook.active

        mapping = self.get_column_mapping(worksheet[1])
        data_list = self.preprare_serializer_data(worksheet, mapping)
        record_valid_data = self.validate_many_serializer(StateDataSerializer, data_list)
        for data in record_valid_data:
            created, state_obj = self.update_or_create_state(state=data["name"], abb=data["abbreviation"])
            result.append({"id": state_obj.id, "name": state_obj.name.title(), "created": created if created else False, "modifed": True if not created else False})
        return Response({"status": 1, "result": result})
    

   

    