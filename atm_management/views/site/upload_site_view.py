from atm_management.generic_view_mixins.custom_api_view import CustomAPIView
from rest_framework.response import Response
from atm_management.serializers import UploadFileSerializer, SiteDataSerializer
from atm_management.exceptions.rest_api_exception import RestAPIException
from openpyxl import load_workbook
from io import BytesIO
from atm_management.models import ATMSite, ATMCity, ATMState
from atm_management.model_queries import ATMSiteModelQueries, ATMStateModelQueries, ATMCityModelQueries
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework import generics  


COLUMN_HEADERS = [
    "Site ID",
    "Site Name",
    "City",
    "State",
    "Pincode",
    "Email",
    "Phone Number",
    "Address Line 1",
    "Address Line 2"
]
COULMNS_LEN = 9

FIELD_COULMN_MAPPING = {
    "Site ID": "site_id",
    "Site Name": "name",
    "City": "city",
    "State": "state",
    "Pincode": "pincode",
    "Email": "email",
    "Phone Number": "phone_number",
    "Address Line 1": "address_line_1",
    "Address Line 2": "address_line_2"
}


class UploadSiteView(CustomAPIView, ATMSiteModelQueries, ATMCityModelQueries, ATMStateModelQueries):
    
    serializer_class = UploadFileSerializer

    def get_column_mapping(self, row):
        mapping_dict = dict()
        for index, cell in enumerate(row):
            if cell.value in COLUMN_HEADERS:
                mapping_dict[FIELD_COULMN_MAPPING[cell.value]] = index 
                
            else:
                raise RestAPIException(f"Incorrect Column Name {cell.value}")

        if len(mapping_dict) < COULMNS_LEN:
            raise RestAPIException(f"Missing Columns from : {COLUMN_HEADERS}")
        return mapping_dict
    
    def preprare_serializer_data(self, worksheet, mapping_dict):
        data_list = list()
        for row in worksheet.iter_rows(min_row=2,max_col=COULMNS_LEN):
            if any(cell.value for cell in row):
                serializer_dict = dict()
                for field in mapping_dict:
                    serializer_dict[field] = row[mapping_dict[field]].value

                data_list.append(serializer_dict)
        return data_list

    def post(self, request):
        result = list()
        req_valid_data = self.validate_serializer(UploadFileSerializer, request.data)      
        workbook = load_workbook(req_valid_data["file"])
        worksheet = workbook.active

        mapping = self.get_column_mapping(worksheet[1])
        data_list = self.preprare_serializer_data(worksheet, mapping)
        record_valid_data = self.validate_many_serializer(SiteDataSerializer, data_list)

        for data in record_valid_data:
            _, state = self.get_or_create_state(data["state"])
            _, city = self.get_or_create_city(data["city"], state.id)
            created, atm_obj = self.update_or_create_atm_site(data, city.id)
            result.append({"id": atm_obj.id, "site_id": atm_obj.site_id, "created": created if created else False, "modifed": True if not created else False})
        return Response({"status": 1, "result": result})
    

   

    def update_or_create_atm_site(self, atm_data, city):
        
        site_id = int(atm_data["site_id"])
        created, atm_obj = self.get_or_create_atm_site(site_id)
        atm_obj.name = atm_data["name"]
        atm_obj.contact_details = {
            "phone_number": int(atm_data["phone_number"]),
            "email": atm_data["email"]
        }
        atm_obj.pincode = atm_data["pincode"]
        atm_obj.address_line_1 = atm_data["address_line_1"]
        atm_obj.address_line_1 = atm_data["address_line_2"]
        atm_obj.city_id = city
        atm_obj.save()
        return created, atm_obj



